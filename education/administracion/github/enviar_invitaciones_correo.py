#!/usr/bin/env python3
"""
Envía invitaciones por correo electrónico institucional (@unimilitar.edu.co)
a la organización umng-mecatronica-ros y asigna a cada estudiante a su respectivo equipo.
"""

import openpyxl
import json
import subprocess
import sys
from pathlib import Path

ORG = "umng-mecatronica-ros"
EXCEL_PATH = Path("education/administracion/github/datos_privados/Lista de Estudiantes ROS.xlsx")

def get_team_id_map():
    cmd = ["gh", "api", f"orgs/{ORG}/teams"]
    res = subprocess.run(cmd, capture_output=True, text=True)
    if res.returncode != 0:
        print("Error obteniendo teams:", res.stderr)
        sys.exit(1)
    teams = json.loads(res.stdout)
    return {t["slug"]: t["id"] for t in teams}

def ensure_team_repo_push_permission(team_slug):
    repo = f"burger-kinova-{team_slug}"
    cmd = [
        "gh", "api", "-X", "PUT",
        f"orgs/{ORG}/teams/{team_slug}/repos/{ORG}/{repo}",
        "-f", "permission=push"
    ]
    res = subprocess.run(cmd, capture_output=True, text=True)
    if res.returncode == 0:
        print(f"  ✓ Repositorio {repo} vinculado a {team_slug} con permiso 'push'")
    else:
        print(f"  ⚠️ Advertencia vinculando {repo} a {team_slug}: {res.stderr.strip()}")

def main():
    if not EXCEL_PATH.is_file():
        print(f"No existe el archivo: {EXCEL_PATH}")
        sys.exit(1)

    print(f"Cargando {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["UniReport_1"]

    team_map = get_team_id_map()
    print("Teams detectados en GitHub:", team_map)

    # Asegurar permisos push para equipos 1 a 6
    print("\n--- Verificando / Asignando permisos push en repositorios ---")
    for i in range(1, 7):
        slug = f"equipo-{i:02d}"
        ensure_team_repo_push_permission(slug)

    print("\n--- Enviando invitaciones por correo institucional ---")
    invitados = 0
    errores = 0

    for row in range(2, ws.max_row + 1):
        codigo = str(ws.cell(row, 1).value or "").strip()
        apellidos = str(ws.cell(row, 2).value or "").strip()
        nombres = str(ws.cell(row, 3).value or "").strip()
        email = str(ws.cell(row, 4).value or "").strip().lower()
        equipo_num = ws.cell(row, 8).value

        if not email or "@" not in email:
            continue

        try:
            equipo_int = int(equipo_num)
            team_slug = f"equipo-{equipo_int:02d}"
        except (ValueError, TypeError):
            print(f"⚠️ Fila {row}: Equipo inválido '{equipo_num}' para {nombres} {apellidos}")
            errores += 1
            continue

        team_id = team_map.get(team_slug)
        if not team_id:
            print(f"❌ No existe el team '{team_slug}' en GitHub para {email}")
            errores += 1
            continue

        payload = {
            "email": email,
            "role": "direct_member",
            "team_ids": [team_id]
        }

        cmd = [
            "gh", "api", "-X", "POST",
            f"orgs/{ORG}/invitations",
            "--input", "-"
        ]

        res = subprocess.run(cmd, input=json.dumps(payload), text=True, capture_output=True)
        if res.returncode == 0:
            print(f"  ✓ Invitación enviada a {email} ({nombres} {apellidos}) $\to$ {team_slug}")
            invitados += 1
        else:
            err = res.stderr.strip().replace("\n", " ")
            if "already a member" in err or "already invited" in err or "422" in err:
                print(f"  ℹ️ {email} ya estaba invitado o es miembro ({team_slug})")
                invitados += 1
            else:
                print(f"  ❌ Error al invitar a {email} ({team_slug}): {err}")
                errores += 1

    print(f"\n==========================================")
    print(f"Resumen: {invitados} invitaciones procesadas correctamente, {errores} errores.")

if __name__ == "__main__":
    main()
